"""One-shot bulk import of provisional player ratings into the roster.

Reads a `.xlsx` of (Name, Gender, Rating, Notes, Singles) rows — the
"Social_Tennis_Registrants_with_Grading" sheet provided by the team —
and adds any names NOT already in the roster, each tagged with the
`Provisional` flag.

Existing roster entries are NEVER overwritten: a name already in the
roster is left untouched (whatever rating / gender / singles it
already has takes precedence).

Usage::

    py -3 scripts/import_provisional_ratings.py <xlsx-path>
    py -3 scripts/import_provisional_ratings.py <xlsx-path> --dry-run

Runs against whatever Google Sheet `roster.py` resolves to via
`GOOGLE_SHEET_ID` in `.env`. Safe to run multiple times — duplicates
are filtered out on each call.

Also writes the "Provisional" column G header if missing, so older
sheets get bootstrapped on first use.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

# Allow running as `py -3 scripts/import_provisional_ratings.py` from
# the repo root (top-level modules like `roster` aren't on sys.path
# otherwise when the script lives in a subdir).
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl  # noqa: E402

from roster import (  # noqa: E402
    COL_PROVISIONAL,
    Roster,
    VALID_GENDERS,
    VALID_SINGLES,
    normalise_rating,
)


def _read_xlsx(path: Path) -> list[dict]:
    """Return a list of {name, gender, rating, singles} dicts."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h or "").strip().lower() for h in rows[0]]
    # Tolerant column lookup so a re-ordered or differently-cased
    # sheet still imports.
    def col(name: str) -> int | None:
        try:
            return header.index(name.lower())
        except ValueError:
            return None
    # `or` would mishandle index 0; coalesce explicitly.
    i_name = col("Full Name")
    if i_name is None:
        i_name = col("Name")
    i_gender = col("Gender")
    i_rating = col("Rating")
    i_singles = col("Singles")
    if i_name is None:
        raise SystemExit(
            f"input xlsx {path} is missing a 'Full Name' / 'Name' column"
        )
    out: list[dict] = []
    for row in rows[1:]:
        name = str(row[i_name] or "").strip()
        if not name:
            continue
        gender = (
            str(row[i_gender] or "").strip().upper()
            if i_gender is not None else ""
        )
        if gender not in VALID_GENDERS:
            gender = ""  # let Roster.add guess
        rating_raw = row[i_rating] if i_rating is not None else None
        try:
            rating = normalise_rating(rating_raw)
        except ValueError:
            rating = "?"
        singles = (
            str(row[i_singles] or "").strip().lower()
            if i_singles is not None else ""
        )
        if singles not in VALID_SINGLES:
            singles = ""
        out.append({
            "name": name, "gender": gender,
            "rating": rating, "singles": singles,
        })
    return out


def _ensure_provisional_header(roster: Roster) -> None:
    """Add the 'Provisional' header at column G if the cell is empty."""
    ws = roster._ws
    current = (ws.cell(1, COL_PROVISIONAL).value or "").strip()
    if current.lower() != "provisional":
        if current and current.lower() != "provisional":
            print(
                f"  warn: column G header is {current!r} — leaving alone "
                "(import will still write data into column G)"
            )
            return
        ws.update_cell(1, COL_PROVISIONAL, "Provisional")
        print("  -> wrote 'Provisional' header to column G")


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("xlsx", type=Path, help="path to the registrants xlsx")
    ap.add_argument(
        "--dry-run", action="store_true",
        help="show what WOULD be added, don't touch the sheet",
    )
    args = ap.parse_args()

    if not args.xlsx.exists():
        raise SystemExit(f"file not found: {args.xlsx}")

    rows = _read_xlsx(args.xlsx)
    print(f"Read {len(rows)} rows from {args.xlsx}")

    roster = Roster()
    existing = set(roster.all())
    print(f"Roster has {len(existing)} existing entries")

    to_add: list[dict] = []
    skipped_existing: list[str] = []
    for r in rows:
        if r["name"] in existing:
            skipped_existing.append(r["name"])
            continue
        to_add.append(r)

    print(f"  {len(skipped_existing)} already in roster -> skipped")
    print(f"  {len(to_add)} will be added as provisional")

    if args.dry_run:
        print("\nDry-run — no changes. First 10 candidates:")
        for r in to_add[:10]:
            print(f"  + {r['name']:<28}  {r['gender']:>1}  rating={r['rating']:>2}  singles={r['singles'] or '-'}")
        return

    # Real run.
    _ensure_provisional_header(roster)
    added_ok = 0
    failed: list[tuple[str, str]] = []
    for r in to_add:
        try:
            roster.add(
                r["name"],
                gender=r["gender"] or None,
                rating=r["rating"],
                singles=r["singles"],
                provisional=True,
            )
            added_ok += 1
            if added_ok % 25 == 0:
                print(f"  ... {added_ok}/{len(to_add)}")
        except Exception as e:
            failed.append((r["name"], str(e)))
    print(f"\nDone. Added {added_ok}. Failed {len(failed)}.")
    for name, err in failed:
        print(f"  FAIL {name!r}: {err}")


if __name__ == "__main__":
    main()
